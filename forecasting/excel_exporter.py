"""Enterprise-grade Excel report builder (4-layer architecture).

Layer 0 — Navigation          00_ReadMe
Layer 1 — Executive           01–05
Layer 2 — Operational         10–17
Layer 3 — Forecast Technical  90–93
Layer 4 — Raw Data            95–99

Does NOT modify forecasting, training, selection, backtest, guardrail,
or DB query logic.  Reuses existing data loaders only.
"""

from __future__ import annotations

import calendar as _cal
import logging
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Literal

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

EXPORT_VERSION = "4.0.0"

# ── Style constants ──────────────────────────────────────────────

_HDR_FONT = Font(bold=True, size=11)
_HDR_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_HDR_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)

_FMT_CHF = "#,##0"
_FMT_PCT = "0.0%"
_FMT_INT = "#,##0"

_TXT_MIN, _TXT_MAX = 18, 45
_NUM_W = 16

# ── Sheet ordering (97_Extra_* slots inserted dynamically) ───────

SHEET_ORDER_FIXED = [
    "00_ReadMe",
    "01_Overview",
    "02_Monthly_Performance",
    "03_Forecast_Summary",
    "04_Risks_And_Actions",
    "05_Strategy",
    "10_Revenue",
    "11_Bookings",
    "12_Customers",
    "13_Utilization",
    "14_Unit_Economics",
    "15_Conversion",
    "16_Stop_Check",
    "17_Notifications",
    "90_Forecast_Daily",
    "91_Fact_vs_Forecast",
    "92_Model_Diagnostics",
    "93_Backtest_Summary",
    "95_Daily_Data",
    "96_Hourly_Data",
]

SHEET_ORDER = SHEET_ORDER_FIXED  # alias for test imports

# ── Legacy dataset key → target sheet mapping ────────────────────

LEGACY_MAP: dict[str, str] = {
    "overview":            "01_Overview",
    "monthly_performance": "02_Monthly_Performance",
    "forecast_summary":    "03_Forecast_Summary",
    "risks":               "04_Risks_And_Actions",
    "revenue_analysis":    "10_Revenue",
    "bookings_analysis":   "11_Bookings",
    "customers_overview":  "12_Customers",
    "utilization":         "13_Utilization",
    "unit_economics":      "14_Unit_Economics",
    "raw_forecasts_daily": "90_Forecast_Daily",
    "fact_vs_forecast":    "91_Fact_vs_Forecast",
    "model_info":          "92_Model_Diagnostics",
    "backtest_summary":    "93_Backtest_Summary",
    "raw_daily":           "95_Daily_Data",
    "raw_hourly":          "96_Hourly_Data",
}

_INTERNAL_KEYS = frozenset({"_daily_frames", "_act_rev", "_act_bk"})


# ═══════════════════════════════════════════════════════════════════
#  Sanitisation
# ═══════════════════════════════════════════════════════════════════

def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Copy, drop Unnamed columns, reset index, ensure unique string headers."""
    df = df.copy()
    df = df.reset_index(drop=True)
    df.columns = [str(c) for c in df.columns]
    bad = [c for c in df.columns if c.startswith("Unnamed")]
    if bad:
        df = df.drop(columns=bad)
    return df


def ensure_no_unnamed(df: pd.DataFrame) -> None:
    bad = [c for c in df.columns if str(c).startswith("Unnamed")]
    assert not bad, f"Unnamed columns found: {bad}"


def _no_data_df() -> pd.DataFrame:
    return pd.DataFrame({"status": ["no_data"]})


# ═══════════════════════════════════════════════════════════════════
#  Sheet writer
# ═══════════════════════════════════════════════════════════════════

def write_table_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    style: Literal["management", "operational", "technical", "raw", "meta"] = "operational",
    currency_cols: list[str] | None = None,
    percent_cols: list[str] | None = None,
    freeze_panes: str = "A2",
    autofilter: bool = True,
) -> None:
    """Write *df* to a new worksheet with category-appropriate styling."""
    df = sanitize_dataframe(df)
    ensure_no_unnamed(df)

    currency_cols = currency_cols or []
    percent_cols = percent_cols or []

    ws = wb.create_sheet(title=sheet_name[:31])
    is_wo = getattr(wb, "write_only", False)

    headers = list(df.columns)
    n_cols = len(headers)
    n_rows = len(df)

    fmt_map: dict[int, str] = {}
    for ci, h in enumerate(headers):
        if h in currency_cols:
            fmt_map[ci] = _FMT_CHF
        elif h in percent_cols:
            fmt_map[ci] = _FMT_PCT

    do_style = style in ("management", "operational")

    if is_wo:
        hdr_cells = []
        for h in headers:
            c = WriteOnlyCell(ws, value=h)
            if do_style:
                c.font = _HDR_FONT
                c.fill = _HDR_FILL
                c.alignment = _HDR_ALIGN
            hdr_cells.append(c)
        ws.append(hdr_cells)

        data = _df_to_native(df)
        apply_fmt = bool(fmt_map) and n_rows <= 5000
        for row_vals in data:
            if apply_fmt:
                cells = []
                for ci, val in enumerate(row_vals):
                    c = WriteOnlyCell(ws, value=val)
                    if ci in fmt_map:
                        c.number_format = fmt_map[ci]
                    cells.append(c)
                ws.append(cells)
            else:
                ws.append(row_vals)
    else:
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            if do_style:
                cell.font = _HDR_FONT
                cell.fill = _HDR_FILL
                cell.alignment = _HDR_ALIGN
        data = _df_to_native(df)
        for row_vals in data:
            ws.append(row_vals)
        if fmt_map and n_rows <= 5000:
            for ci, fmt in fmt_map.items():
                col_idx = ci + 1
                for r in range(2, n_rows + 2):
                    ws.cell(row=r, column=col_idx).number_format = fmt

    if freeze_panes:
        ws.freeze_panes = freeze_panes
    if autofilter and n_rows > 0 and style != "meta":
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"

    _auto_widths(ws, headers, df)


# ═══════════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════════

def _df_to_native(df: pd.DataFrame) -> list[list]:
    n = len(df)
    if n == 0:
        return []
    cols: list[list] = []
    for col in df.columns:
        s = df[col]
        k = s.dtype.kind
        if k == "f":
            raw = s.values
            mask = np.isfinite(raw)
            native = raw.tolist()
            for i in range(n):
                if not mask[i]:
                    native[i] = None
            cols.append(native)
        elif k in ("i", "u"):
            cols.append(s.astype(int).tolist())
        elif k == "b":
            cols.append(s.astype(bool).tolist())
        else:
            native = s.tolist()
            for i in range(n):
                v = native[i]
                if v is pd.NaT or (isinstance(v, float) and not np.isfinite(v)):
                    native[i] = None
            cols.append(native)
    return [list(row) for row in zip(*cols)]


def _auto_widths(ws, headers, df):
    sample = df.head(200) if len(df) > 200 else df
    for ci, col in enumerate(headers, start=1):
        hl = len(str(col))
        if sample[col].dtype in ("object", "string"):
            mx = sample[col].astype(str).str.len().max() if len(sample) > 0 else 0
            w = min(max(hl, int(mx) if pd.notna(mx) else 0, _TXT_MIN), _TXT_MAX)
        else:
            w = max(hl + 2, _NUM_W)
        ws.column_dimensions[get_column_letter(ci)].width = w


def _safe_div(a, b, default=None):
    if b is None or b == 0:
        return default
    return a / b


def _bt_metric(m: str) -> str:
    from config import METRIC_OPTIMIZATION
    return METRIC_OPTIMIZATION.get(m, "smape")


def _latest_trained(metrics, load_spec_fn) -> str:
    latest = ""
    for m in metrics:
        spec = load_spec_fn(m)
        if spec and spec.get("trained_at", "") > latest:
            latest = spec["trained_at"]
    return latest or "n/a"


# ═══════════════════════════════════════════════════════════════════
#  Data collection (existing loaders ONLY)
# ═══════════════════════════════════════════════════════════════════

def collect_export_datasets(
    date_from: date | None = None,
    date_to: date | None = None,
    year: int | None = None,
    month: int | None = None,
    token: str | None = None,
    tenant_id: int | None = None,
) -> tuple[dict[str, pd.DataFrame], dict[str, str]]:
    """Gather data from existing loaders.  Returns (datasets, metadata)."""
    from config import METRICS, DEFAULT_HORIZON_DAYS
    from repository import (
        load_metric, load_latest_forecast, load_spec,
        load_daily_revenue_by_activity, load_daily_bookings_by_activity,
        list_segments, list_monthly_rollups,
    )
    from registry import get_leaderboard, _load_runs

    now = datetime.now()
    since = date_from or (now.date() - timedelta(days=400))
    until = date_to or now.date()
    if year is None:
        year = now.year
    if month is None:
        month = now.month

    ds: dict[str, pd.DataFrame] = {}

    # ── Daily metrics ────────────────────────────────
    daily_frames: dict[str, pd.DataFrame] = {}
    for metric in METRICS:
        try:
            ts = load_metric(metric, token=token)
            mask = (ts.dates.date >= since) & (ts.dates.date <= until)
            daily_frames[metric] = pd.DataFrame({
                "date": ts.dates[mask].strftime("%Y-%m-%d"),
                metric: ts.values[mask],
            })
        except Exception as e:
            logger.warning("load %s failed: %s", metric, e)

    if daily_frames:
        base = None
        for m, df in daily_frames.items():
            base = df if base is None else base.merge(df, on="date", how="outer")
        if base is not None:
            ds["raw_daily"] = base.sort_values("date").reset_index(drop=True)

    # ── Forecasts ────────────────────────────────────
    fc_results: dict[str, Any] = {}
    for metric in METRICS:
        try:
            fc = load_latest_forecast(metric, tenant_id=tenant_id)
            if fc:
                fc_results[metric] = fc
        except Exception:
            pass

    fc_daily_rows: list[dict] = []
    for metric in METRICS:
        fc = fc_results.get(metric)
        if not fc:
            continue
        gen_at = fc.created_at or ""
        for i, d in enumerate(fc.dates):
            fc_daily_rows.append({
                "metric": metric, "date": d,
                "yhat": fc.yhat[i],
                "yhat_lower": fc.yhat_lower[i] if fc.yhat_lower is not None else None,
                "yhat_upper": fc.yhat_upper[i] if fc.yhat_upper is not None else None,
                "model_family": fc.model_family or fc.model_name,
                "generated_at": gen_at, "confidence_level": 0.80,
            })
    ds["raw_forecasts_daily"] = pd.DataFrame(fc_daily_rows) if fc_daily_rows else _no_data_df()

    # ── Activity breakdown ───────────────────────────
    seg_list = list_segments()
    seg_names = {s["id"]: s["name"] for s in seg_list}

    act_rev: dict[str, float] = {}
    act_bk: dict[str, float] = {}
    for label, loader, out_dict in [
        ("revenue", load_daily_revenue_by_activity, act_rev),
        ("bookings", load_daily_bookings_by_activity, act_bk),
    ]:
        try:
            by_act = loader(since=since)
            for sid, ts in (by_act or {}).items():
                name = seg_names.get(sid, f"Activity {sid}")
                out_dict[name] = float(np.sum(ts.values))
        except Exception:
            pass

    # ── Monthly rollups ──────────────────────────────
    monthly_rollup_cache: dict[str, list[dict]] = {}
    for metric in ("revenue", "bookings"):
        try:
            monthly_rollup_cache[metric] = list_monthly_rollups(metric, limit=24, tenant_id=tenant_id)
        except Exception:
            monthly_rollup_cache[metric] = []

    # ── Model info (92_Model_Diagnostics) ────────────
    model_rows: list[dict] = []
    for metric in METRICS:
        spec = load_spec(metric)
        lb = get_leaderboard(metric)
        best_score = lb[0].mean_error if lb else None
        model_rows.append({
            "metric": metric,
            "selected_model_family": spec.get("model_family", "") if spec else "",
            "transform": spec.get("transform_name", "") if spec else "",
            "folds": spec.get("diagnostics", {}).get("n_folds", "") if spec else "",
            "horizon_days": DEFAULT_HORIZON_DAYS,
            "backtest_metric": _bt_metric(metric),
            "backtest_score": round(best_score, 6) if best_score is not None else None,
            "residual_test": None,
            "guardrails": "yes",
            "notes": ", ".join(spec.get("warnings", [])) if spec else "",
        })
    ds["model_info"] = pd.DataFrame(model_rows)

    # ── Backtest summary (93) ────────────────────────
    bt_rows: list[dict] = []
    for metric in METRICS:
        spec = load_spec(metric, tenant_id=tenant_id)
        lb = get_leaderboard(metric)
        selected_fam = spec.get("model_family", "") if spec else ""
        for entry in lb:
            fam = getattr(entry, "model_family", getattr(entry, "name", selected_fam))
            bt_rows.append({
                "metric": metric,
                "model_family": fam,
                "folds": spec.get("diagnostics", {}).get("n_folds", "") if spec else "",
                "horizon_days": DEFAULT_HORIZON_DAYS,
                "mean_score": round(entry.mean_error, 6),
                "std_score": getattr(entry, "std_error", None),
                "min_score": getattr(entry, "min_error", None),
                "max_score": getattr(entry, "max_error", None),
                "selected": (fam == selected_fam),
            })
    ds["backtest_summary"] = pd.DataFrame(bt_rows) if bt_rows else _no_data_df()

    # ── Fact vs Forecast (91) ────────────────────────
    fvf_rows: list[dict] = []
    raw_d = ds.get("raw_daily")
    if raw_d is not None and "status" not in raw_d.columns:
        for metric in METRICS:
            if metric not in raw_d.columns:
                continue
            fc = fc_results.get(metric)
            if not fc:
                continue
            fc_date_strs = pd.to_datetime(fc.dates).strftime("%Y-%m-%d")
            fc_map = dict(zip(fc_date_strs, fc.yhat))
            fc_date_set = set(fc_date_strs)
            actuals = raw_d[["date", metric]].dropna(subset=[metric])
            relevant = actuals[actuals["date"].isin(fc_date_set)]
            for _, row in relevant.iterrows():
                d = row["date"]
                actual = float(row[metric])
                yhat = float(fc_map[d])
                err_abs = abs(actual - yhat)
                err_pct = (err_abs / abs(actual) * 100) if abs(actual) > 1e-9 else None
                denom = abs(actual) + abs(yhat)
                smape_v = (2 * err_abs / denom * 100) if denom > 1e-9 else None
                fvf_rows.append({
                    "metric": metric, "date": d,
                    "actual": round(actual, 2), "yhat": round(yhat, 2),
                    "error_abs": round(err_abs, 2),
                    "error_pct": round(err_pct, 2) if err_pct is not None else None,
                    "smape": round(smape_v, 2) if smape_v is not None else None,
                })
    ds["fact_vs_forecast"] = pd.DataFrame(fvf_rows) if fvf_rows else _no_data_df()

    # ── Risks (04) ───────────────────────────────────
    risk_rows: list[dict] = []
    for metric in METRICS:
        df = daily_frames.get(metric)
        if df is None or len(df) < 30:
            continue
        vals = df[metric].dropna().values.astype(float)
        if len(vals) < 30:
            continue
        mu, sigma = np.mean(vals), np.std(vals)
        if sigma < 1e-9:
            continue
        z = (vals - mu) / sigma
        dates_arr = df["date"].values
        for i, zv in enumerate(z):
            if abs(zv) >= 2.5:
                direction = "spike" if zv > 0 else "drop"
                risk_rows.append({
                    "Type": "Risk" if direction == "drop" else "Opportunity",
                    "Severity": round(min(abs(zv) * 20, 100)),
                    "Title": f"{metric} {direction} on {dates_arr[i]}",
                    "Metric Evidence": metric,
                    "Evidence Value": f"z={zv:.1f} (val={vals[i]:.0f}, avg={mu:.0f})",
                    "Recommended Action": f"Investigate {metric} {direction}",
                    "_sort": abs(zv),
                })
    if risk_rows:
        risk_df = pd.DataFrame(risk_rows).sort_values("_sort", ascending=False)
        if len(risk_df) > 10:
            extra_n = len(risk_df) - 9
            risk_df = risk_df.head(9)
            risk_df = pd.concat([risk_df, pd.DataFrame([{
                "Type": "Info", "Severity": 0,
                "Title": f"\u2026 {extra_n} more items",
                "Metric Evidence": "", "Evidence Value": "",
                "Recommended Action": "", "_sort": 0,
            }])], ignore_index=True)
        risk_df = risk_df.drop(columns=["_sort"])
        ds["risks"] = risk_df
    else:
        ds["risks"] = pd.DataFrame([{
            "Type": "Info", "Severity": 0,
            "Title": "No risks detected",
            "Metric Evidence": "", "Evidence Value": "",
            "Recommended Action": "",
        }])

    # ── Operational tables ───────────────────────────
    rd = daily_frames.get("revenue")
    bd = daily_frames.get("bookings")
    cd = daily_frames.get("cancel_rate")
    ud = daily_frames.get("utilization")

    if rd is not None and len(rd) > 0:
        ra = rd.rename(columns={"revenue": "Revenue", "date": "Date"}).copy()
        ra["Revenue 7D Avg"] = ra["Revenue"].rolling(7, min_periods=1).mean().round(2)
        ra["Revenue vs Prev Week %"] = ra["Revenue"].pct_change(7)
        ds["revenue_analysis"] = ra
    else:
        ds["revenue_analysis"] = _no_data_df()

    if bd is not None and len(bd) > 0:
        ba = bd.rename(columns={"bookings": "Bookings", "date": "Date"}).copy()
        ba["Bookings 7D Avg"] = ba["Bookings"].rolling(7, min_periods=1).mean().round(2)
        ba["Bookings vs Prev Week %"] = ba["Bookings"].pct_change(7)
        ds["bookings_analysis"] = ba
    else:
        ds["bookings_analysis"] = _no_data_df()

    ds["customers_overview"] = _no_data_df()

    if ud is not None and len(ud) > 0:
        u = ud.rename(columns={"date": "Date", "utilization": "Utilization %"}).copy()
        u["Utilization %"] = u["Utilization %"] / 100  # store 0-1 for percent format
        ds["utilization"] = u
    else:
        ds["utilization"] = _no_data_df()

    activities = sorted(set(list(act_rev.keys()) + list(act_bk.keys())))
    if activities:
        ue_rows = []
        for a in activities:
            rev = act_rev.get(a, 0)
            bk = act_bk.get(a, 0)
            ue_rows.append({
                "Activity": a,
                "Revenue": round(rev, 2),
                "Bookings": round(bk, 0),
                "Avg Check": round(rev / max(bk, 1), 2) if bk > 0 else None,
                "RevPAH": None,
                "Utilization %": None,
                "Cancel Rate %": None,
            })
        ds["unit_economics"] = pd.DataFrame(ue_rows)
    else:
        ds["unit_economics"] = _no_data_df()

    # ── Management tables ────────────────────────────
    # Next month
    nm_year, nm_month = (year + 1, 1) if month == 12 else (year, month + 1)
    total_days_nm = _cal.monthrange(nm_year, nm_month)[1]
    first_nm = pd.Timestamp(year=nm_year, month=nm_month, day=1)
    last_nm = pd.Timestamp(year=nm_year, month=nm_month, day=total_days_nm)

    rev_total = float(rd["revenue"].sum()) if rd is not None and len(rd) > 0 else None
    bk_total = float(bd["bookings"].sum()) if bd is not None and len(bd) > 0 else None
    avg_check = _safe_div(rev_total, bk_total)
    cancel_avg = float(cd["cancel_rate"].mean()) if cd is not None and len(cd) > 0 else None
    util_avg = float(ud["utilization"].mean()) if ud is not None and len(ud) > 0 else None

    top_activity = max(act_rev, key=act_rev.get) if act_rev else None

    rev_vs_prev = None
    bk_vs_prev = None
    if rd is not None and len(rd) >= 14:
        half = len(rd) // 2
        r1, r2 = rd["revenue"].iloc[:half].sum(), rd["revenue"].iloc[half:].sum()
        if r1 > 0:
            rev_vs_prev = round((r2 - r1) / r1, 4)
    if bd is not None and len(bd) >= 14:
        half = len(bd) // 2
        b1, b2 = bd["bookings"].iloc[:half].sum(), bd["bookings"].iloc[half:].sum()
        if b1 > 0:
            bk_vs_prev = round((b2 - b1) / b1, 4)

    # Build forecast totals for next month (shared between overview & forecast_summary)
    fc_rev_total = fc_rev_lo = fc_rev_hi = None
    fc_bk_total = fc_bk_lo = fc_bk_hi = None
    fc_model = ""
    covered_days = 0

    for mkey, prefix_pair in [("revenue", "rev"), ("bookings", "bk")]:
        got = False
        for r in monthly_rollup_cache.get(mkey, []):
            if r.get("year") == nm_year and r.get("month") == nm_month:
                if mkey == "revenue":
                    fc_rev_total = r.get("predicted_total")
                    fc_rev_lo = r.get("lower_total")
                    fc_rev_hi = r.get("upper_total")
                    fc_model = r.get("model_family_used", "")
                else:
                    fc_bk_total = r.get("predicted_total")
                    fc_bk_lo = r.get("lower_total")
                    fc_bk_hi = r.get("upper_total")
                got = True
                break
        if not got:
            fc = fc_results.get(mkey)
            if fc:
                fc_dates = pd.to_datetime(fc.dates)
                mask = (fc_dates >= first_nm) & (fc_dates <= last_nm)
                if mask.any():
                    yhat = np.array(fc.yhat)[mask]
                    total = round(float(yhat.sum()), 2)
                    lo = round(float(np.array(fc.yhat_lower)[mask].sum()), 2) if fc.yhat_lower is not None else None
                    hi = round(float(np.array(fc.yhat_upper)[mask].sum()), 2) if fc.yhat_upper is not None else None
                    if mkey == "revenue":
                        fc_rev_total, fc_rev_lo, fc_rev_hi = total, lo, hi
                        fc_model = fc.model_family or fc.model_name or ""
                        covered_days = int(mask.sum())
                    else:
                        fc_bk_total, fc_bk_lo, fc_bk_hi = total, lo, hi

    if covered_days == 0:
        fc = fc_results.get("revenue")
        if fc:
            fc_dates = pd.to_datetime(fc.dates)
            covered_days = int(((fc_dates >= first_nm) & (fc_dates <= last_nm)).sum())

    cov_status = "no_data" if covered_days == 0 else ("partial" if covered_days < total_days_nm else "full")
    cov_ratio = round(covered_days / total_days_nm, 4) if total_days_nm > 0 else 0

    # Forecast for extra metrics (mean aggregation over next month)
    fc_cancel = fc_util = fc_avgchk = None
    for mkey, is_pct in [("cancel_rate", True), ("utilization", True), ("avg_check", False)]:
        fc = fc_results.get(mkey)
        if not fc:
            continue
        fc_dates = pd.to_datetime(fc.dates)
        mask = (fc_dates >= first_nm) & (fc_dates <= last_nm)
        if not mask.any():
            continue
        vals = np.array(fc.yhat)[mask]
        if mkey == "avg_check":
            bk_fc = fc_results.get("bookings")
            if bk_fc:
                bk_dates = pd.to_datetime(bk_fc.dates)
                bk_mask = (bk_dates >= first_nm) & (bk_dates <= last_nm)
                if bk_mask.any():
                    bk_v = np.array(bk_fc.yhat)[bk_mask]
                    min_len = min(len(vals), len(bk_v))
                    tot = bk_v[:min_len].sum()
                    fc_avgchk = round(float((vals[:min_len] * bk_v[:min_len]).sum() / max(tot, 1)), 2) if tot > 0 else round(float(vals.mean()), 2)
                else:
                    fc_avgchk = round(float(vals.mean()), 2)
            else:
                fc_avgchk = round(float(vals.mean()), 2)
        elif mkey == "cancel_rate":
            fc_cancel = round(float(vals.mean()), 4)
        elif mkey == "utilization":
            fc_util = round(float(vals.mean()) / 100, 4)  # 0-1

    # Historical averages
    hist_rev = hist_bk = None
    if rd is not None and len(rd) > 0:
        rdf = rd.copy()
        rdf["_dt"] = pd.to_datetime(rdf["date"])
        sm = rdf[rdf["_dt"].dt.month == nm_month]
        if len(sm) > 0:
            hist_rev = round(float(sm["revenue"].sum()) / max(sm["_dt"].dt.year.nunique(), 1), 2)
    if bd is not None and len(bd) > 0:
        bdf = bd.copy()
        bdf["_dt"] = pd.to_datetime(bdf["date"])
        sm = bdf[bdf["_dt"].dt.month == nm_month]
        if len(sm) > 0:
            hist_bk = round(float(sm["bookings"].sum()) / max(sm["_dt"].dt.year.nunique(), 1), 2)

    fc_vs_hist = None
    if fc_rev_total and hist_rev and hist_rev > 0:
        fc_vs_hist = round((fc_rev_total - hist_rev) / hist_rev, 4)

    # 01_Overview (20 columns, 1 row)
    ds["overview"] = pd.DataFrame([{
        "Period From": str(since),
        "Period To": str(until),
        "Revenue Total": round(rev_total, 2) if rev_total else None,
        "Bookings Total": round(bk_total, 0) if bk_total else None,
        "Avg Check": round(avg_check, 2) if avg_check else None,
        "Cancel Rate %": round(cancel_avg, 4) if cancel_avg is not None else None,
        "Utilization %": round(util_avg / 100, 4) if util_avg is not None else None,
        "Revenue vs Prev Period %": rev_vs_prev,
        "Bookings vs Prev Period %": bk_vs_prev,
        "Top Activity (by Revenue)": top_activity,
        "Biggest Growth Area (MoM Revenue)": None,
        "Forecast Month (YYYY-MM)": f"{nm_year}-{nm_month:02d}",
        "Forecast Revenue (Month Total)": fc_rev_total,
        "Forecast Revenue CI Low": fc_rev_lo,
        "Forecast Revenue CI High": fc_rev_hi,
        "Forecast Bookings (Month Total)": fc_bk_total,
        "Forecast Bookings CI Low": fc_bk_lo,
        "Forecast Bookings CI High": fc_bk_hi,
        "Model Family Used (Revenue)": fc_model,
        "Notes": "",
    }])

    # 02_Monthly_Performance
    monthly_rows: list[dict] = []
    if rd is not None and len(rd) > 0:
        rdf = rd.copy()
        rdf["_m"] = pd.to_datetime(rdf["date"]).dt.to_period("M")
        grouped: dict[str, dict[str, float]] = {}
        for period, grp in rdf.groupby("_m"):
            grouped[str(period)] = {"revenue": float(grp["revenue"].sum())}
        for metric_df, mkey in [(bd, "bookings"), (cd, "cancel_rate"), (ud, "utilization")]:
            if metric_df is None:
                continue
            tmp = metric_df.copy()
            tmp["_m"] = pd.to_datetime(tmp["date"]).dt.to_period("M")
            agg = "mean" if mkey in ("cancel_rate", "utilization") else "sum"
            for period, grp in tmp.groupby("_m"):
                v = float(grp[mkey].mean()) if agg == "mean" else float(grp[mkey].sum())
                grouped.setdefault(str(period), {})[mkey] = v

        prev_r = None
        for ym in sorted(grouped):
            v = grouped[ym]
            rev = v.get("revenue")
            bk = v.get("bookings")
            util = v.get("utilization")
            cr = v.get("cancel_rate")
            row = {
                "Month": ym,
                "Revenue": round(rev, 2) if rev else None,
                "Bookings": round(bk, 0) if bk else None,
                "Avg Check": round(rev / max(bk, 1), 2) if rev and bk else None,
                "MoM Growth %": None,
                "Utilization %": round(util / 100, 4) if util is not None else None,
                "Cancel Rate %": round(cr, 4) if cr is not None else None,
            }
            if prev_r is not None and prev_r > 0 and rev is not None:
                row["MoM Growth %"] = round((rev - prev_r) / prev_r, 4)
            prev_r = rev
            monthly_rows.append(row)
    ds["monthly_performance"] = pd.DataFrame(monthly_rows) if monthly_rows else _no_data_df()

    # 03_Forecast_Summary (17 columns, 1 row)
    ds["forecast_summary"] = pd.DataFrame([{
        "Forecast Month (YYYY-MM)": f"{nm_year}-{nm_month:02d}",
        "Forecast Revenue (Total)": fc_rev_total,
        "Forecast Revenue CI Low": fc_rev_lo,
        "Forecast Revenue CI High": fc_rev_hi,
        "Forecast Bookings (Total)": fc_bk_total,
        "Forecast Bookings CI Low": fc_bk_lo,
        "Forecast Bookings CI High": fc_bk_hi,
        "Forecast Cancel Rate % (Avg)": fc_cancel,
        "Forecast Utilization % (Avg)": fc_util,
        "Forecast Avg Check (Avg / Weighted)": fc_avgchk,
        "Model Family Used (Revenue)": fc_model,
        "Confidence Level": 0.80,
        "Last Updated At (ISO)": _latest_trained(METRICS, lambda m: load_spec(m, tenant_id=tenant_id)),
        "Coverage Status": cov_status,
        "Coverage Ratio": cov_ratio,
        "Historical Avg Revenue (Same Month)": hist_rev,
        "Forecast vs Historical Avg %": fc_vs_hist,
    }])

    # ── Datasets with no loader ──────────────────────
    ds["raw_hourly"] = _no_data_df()
    ds["raw_customers"] = _no_data_df()

    # ── Metadata ─────────────────────────────────────
    metadata = {
        "report_generated_at": now.isoformat(),
        "report_timezone": "Europe/Zurich",
        "date_from": str(since),
        "date_to": str(until),
        "currency": "CHF",
        "restaurant_id": "all",
        "export_version": EXPORT_VERSION,
        "forecast_horizon_days": str(DEFAULT_HORIZON_DAYS),
        "confidence_level": "0.80",
        "data_freshness_note": f"Data as of {now.strftime('%Y-%m-%d %H:%M')}",
    }

    return ds, metadata


# ═══════════════════════════════════════════════════════════════════
#  Layer builders
# ═══════════════════════════════════════════════════════════════════

def build_management_layer(
    datasets: dict[str, pd.DataFrame],
    metadata: dict[str, str],
) -> dict[str, pd.DataFrame]:
    return {
        "00_ReadMe": _build_readme(),
        "01_Overview": datasets.get("overview", _no_data_df()),
        "02_Monthly_Performance": datasets.get("monthly_performance", _no_data_df()),
        "03_Forecast_Summary": datasets.get("forecast_summary", _no_data_df()),
        "04_Risks_And_Actions": datasets.get("risks", _no_data_df()),
        "05_Strategy": _no_data_df(),
    }


def build_operational_layer(
    datasets: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    return {
        "10_Revenue":       datasets.get("revenue_analysis", _no_data_df()),
        "11_Bookings":      datasets.get("bookings_analysis", _no_data_df()),
        "12_Customers":     datasets.get("customers_overview", _no_data_df()),
        "13_Utilization":   datasets.get("utilization", _no_data_df()),
        "14_Unit_Economics": datasets.get("unit_economics", _no_data_df()),
        "15_Conversion":    _no_data_df(),
        "16_Stop_Check":    _no_data_df(),
        "17_Notifications": _no_data_df(),
    }


def build_forecast_technical_layer(
    datasets: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    return {
        "90_Forecast_Daily":   datasets.get("raw_forecasts_daily", _no_data_df()),
        "91_Fact_vs_Forecast": datasets.get("fact_vs_forecast", _no_data_df()),
        "92_Model_Diagnostics": datasets.get("model_info", _no_data_df()),
        "93_Backtest_Summary": datasets.get("backtest_summary", _no_data_df()),
    }


def build_raw_layer(
    datasets: dict[str, pd.DataFrame],
    metadata: dict[str, str],
) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {
        "95_Daily_Data":  datasets.get("raw_daily", _no_data_df()),
        "96_Hourly_Data": datasets.get("raw_hourly", _no_data_df()),
    }

    # 97_Extra_* for any unmapped datasets (legacy preservation)
    mapped_keys = set(LEGACY_MAP.keys())
    used_names: set[str] = set()
    for key, df in sorted(datasets.items()):
        if key in mapped_keys or key in _INTERNAL_KEYS or key.startswith("_"):
            continue
        extra_name = f"97_Extra_{key}"[:31]
        base = extra_name
        counter = 2
        while extra_name in used_names:
            extra_name = f"{base[:28]}_{counter}"
            counter += 1
        used_names.add(extra_name)
        out[extra_name] = df

    out["99_Metadata"] = build_metadata_sheet(metadata)
    return out


def build_metadata_sheet(metadata: dict[str, str]) -> pd.DataFrame:
    return pd.DataFrame([{"key": k, "value": v} for k, v in metadata.items()])


def _build_readme() -> pd.DataFrame:
    rows = [
        {"Section": "Executive (01\u201305)",
         "What you see": "High-level KPIs, monthly trends, forecast summary, risks & strategy",
         "How to use": "Start here for a quick business overview"},
        {"Section": "Operational (10\u201317)",
         "What you see": "Daily revenue/bookings, customers, utilization, unit economics",
         "How to use": "Drill into specific areas for operational decisions"},
        {"Section": "Forecast Technical (90\u201393)",
         "What you see": "Raw forecast output, fact-vs-forecast accuracy, model diagnostics",
         "How to use": "Validate model performance and forecast quality"},
        {"Section": "Raw Data (95\u201399)",
         "What you see": "Complete underlying data tables and metadata",
         "How to use": "Export to BI tools or custom analysis"},
        {"Section": "Key Term: sMAPE",
         "What you see": "Symmetric Mean Absolute Percentage Error (0% = perfect, lower = better)",
         "How to use": "Compare forecast accuracy across metrics"},
        {"Section": "Key Term: Confidence Interval",
         "What you see": "Range where the true value is expected to fall (e.g. 80% CI)",
         "How to use": "Wider intervals indicate higher uncertainty"},
        {"Section": "Key Term: Forecast Month",
         "What you see": "The calendar month being predicted (e.g. 2026-03 = March 2026)",
         "How to use": "Shows the target period for financial projections"},
    ]
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════
#  Workbook builder
# ═══════════════════════════════════════════════════════════════════

_SHEET_SPEC: dict[str, dict] = {
    "00_ReadMe":             {"style": "management"},
    "01_Overview":           {"style": "management",
                              "currency": ["Revenue Total", "Avg Check",
                                           "Forecast Revenue (Month Total)",
                                           "Forecast Revenue CI Low", "Forecast Revenue CI High"],
                              "percent": ["Cancel Rate %", "Utilization %",
                                          "Revenue vs Prev Period %", "Bookings vs Prev Period %"]},
    "02_Monthly_Performance": {"style": "management",
                               "currency": ["Revenue", "Avg Check"],
                               "percent": ["MoM Growth %", "Utilization %", "Cancel Rate %"]},
    "03_Forecast_Summary":   {"style": "management",
                              "currency": ["Forecast Revenue (Total)", "Forecast Revenue CI Low",
                                           "Forecast Revenue CI High",
                                           "Forecast Avg Check (Avg / Weighted)",
                                           "Historical Avg Revenue (Same Month)"],
                              "percent": ["Forecast Cancel Rate % (Avg)", "Forecast Utilization % (Avg)",
                                          "Forecast vs Historical Avg %", "Coverage Ratio"]},
    "04_Risks_And_Actions":  {"style": "management"},
    "05_Strategy":           {"style": "management"},
    "10_Revenue":            {"style": "operational",
                              "currency": ["Revenue", "Revenue 7D Avg"],
                              "percent": ["Revenue vs Prev Week %"]},
    "11_Bookings":           {"style": "operational",
                              "percent": ["Bookings vs Prev Week %"]},
    "12_Customers":          {"style": "operational", "percent": ["Cancel Rate %"]},
    "13_Utilization":        {"style": "operational", "percent": ["Utilization %"]},
    "14_Unit_Economics":     {"style": "operational",
                              "currency": ["Revenue", "Avg Check"],
                              "percent": ["Utilization %", "Cancel Rate %"]},
    "15_Conversion":         {"style": "operational"},
    "16_Stop_Check":         {"style": "operational"},
    "17_Notifications":      {"style": "operational"},
    "90_Forecast_Daily":     {"style": "technical"},
    "91_Fact_vs_Forecast":   {"style": "technical"},
    "92_Model_Diagnostics":  {"style": "technical"},
    "93_Backtest_Summary":   {"style": "technical"},
    "95_Daily_Data":         {"style": "raw"},
    "96_Hourly_Data":        {"style": "raw"},
    "99_Metadata":           {"style": "meta", "autofilter": False},
}


def build_workbook(
    datasets: dict[str, pd.DataFrame],
    metadata: dict[str, str],
    *,
    write_only: bool = True,
) -> Workbook:
    wb = Workbook(write_only=write_only)
    if not write_only:
        wb.remove(wb.active)

    mgmt = build_management_layer(datasets, metadata)
    ops = build_operational_layer(datasets)
    tech = build_forecast_technical_layer(datasets)
    raw = build_raw_layer(datasets, metadata)

    all_sheets: dict[str, pd.DataFrame] = {}
    all_sheets.update(mgmt)
    all_sheets.update(ops)
    all_sheets.update(tech)
    all_sheets.update(raw)

    extra_names = sorted(k for k in all_sheets if k.startswith("97_Extra_"))
    order = list(SHEET_ORDER_FIXED) + extra_names + ["99_Metadata"]

    for name in order:
        df = all_sheets.get(name)
        if df is None or df.empty:
            df = _no_data_df()
        spec = _SHEET_SPEC.get(name, {})
        sty = spec.get("style", "raw" if name.startswith("97_") else "operational")
        write_table_sheet(
            wb, name, df,
            style=sty,
            currency_cols=spec.get("currency", []),
            percent_cols=spec.get("percent", []),
            autofilter=spec.get("autofilter", True),
        )

    return wb


# ═══════════════════════════════════════════════════════════════════
#  Byte serialisation
# ═══════════════════════════════════════════════════════════════════

def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
